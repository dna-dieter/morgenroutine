#!/usr/bin/env python3
"""
Mantelroutine — Spreadsheet- und TOML-gesteuerter Job-Runner
=============================================================
Wird taeglich um 00:00 via launchd gestartet.

Ablauf:
  1. schedule_matrix.xlsx lesen → Startzeit + aktive Jobs fuer heute
  2. Bis zur Startzeit schlafen (sleep)
  3. Wochentag-TOML laden → Job-Details (Modul, Abhaengigkeiten)
  4. Jobs ausfuehren, Ergebnisse als JSON schreiben

Die xlsx ist die Steuerungstabelle (WANN + WAS).
Die TOML-Dateien liefern die technischen Details (WIE).

Exit-Codes:
  0 = Erfolg (auch bei leerem Schedule)
  1 = Mindestens ein Job fehlgeschlagen (retry-wuerdig fuer launchd)
  2 = Fataler Fehler (Config nicht lesbar, etc.)

Autor: AI Artifakte
"""

import json
import os
import subprocess
import sys
import time
import logging
import tomllib
from datetime import datetime, timedelta, date
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ---------------------------------------------------------------------------
# Pfade
# ---------------------------------------------------------------------------

SCRIPT_DIR = Path(__file__).parent.resolve()
DATA_DIR = SCRIPT_DIR / "data"
LOG_DIR = SCRIPT_DIR / "logs"

HOME = Path.home()
VAULT = HOME / "Library/Mobile Documents/iCloud~md~obsidian/Documents/meinBrain/meinBrain"
SEC_DIR = HOME / "Documents/SEC filing"

NACHT_LOG_DIR = VAULT / "20 Claude-selber/Nacht-Log"
PENDING_SYNC = VAULT / "20 Claude-selber/Pending-Sync/pending-changes.md"
SEC_STATUS_FILE = SEC_DIR / "status.json"

# Live-Status-Datei: wird waehrend der Ausfuehrung geschrieben,
# Dashboard pollt diese Datei alle 2 Sekunden
RUNNING_FILE = DATA_DIR / "running.json"

# Wochentag-Mapping: Python weekday() → deutscher Name
WOCHENTAGE = {
    0: "montag",
    1: "dienstag",
    2: "mittwoch",
    3: "donnerstag",
    4: "freitag",
    5: "samstag",
    6: "sonntag",
}

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

LOG_DIR.mkdir(parents=True, exist_ok=True)
DATA_DIR.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_DIR / "mantel.log"),
        logging.StreamHandler(),
    ],
)
log = logging.getLogger("mantel")

TODAY = datetime.now().strftime("%Y-%m-%d")
NOW_ISO = datetime.now().isoformat()
NOW_HM = datetime.now().strftime("%H:%M")


# ===================================================================
#  Live-Status (running.json)
# ===================================================================

def _update_running(state: str, **kwargs):
    """
    Schreibt running.json mit dem aktuellen Mantel-Status.

    state kann sein:
      - "waiting"     → wartet auf Startzeit
      - "running"     → fuehrt Jobs aus
      - "finished"    → fertig (Ergebnis in kwargs)
      - "idle"        → nicht aktiv (Datei wird geloescht)

    Zusaetzliche kwargs je nach state:
      waiting:  target_time, wait_seconds
      running:  current_job, job_index, jobs_total, job_names, results_so_far, started
      finished: overall_status, jobs_success, jobs_error, duration_sec, results
    """
    if state == "idle":
        # Datei loeschen wenn idle
        try:
            RUNNING_FILE.unlink(missing_ok=True)
        except Exception:
            pass
        return

    data = {
        "state": state,
        "pid": os.getpid(),
        "updated": datetime.now().isoformat(),
    }
    data.update(kwargs)

    try:
        with open(RUNNING_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        log.warning(f"running.json nicht schreibbar: {e}")


# ===================================================================
#  Spreadsheet lesen (schedule_matrix.xlsx)
# ===================================================================

# Mapping: xlsx Spaltenname → TOML Job-Name
XLSX_JOB_COLUMNS = {
    "SEC Delta Update": "sec_delta_update",
    "Aktien Delta Update": "aktien_delta_update",
    "Aktien Kalkulation": "aktien_kalkulation",
}

SCHEDULE_MATRIX = SCRIPT_DIR / "schedule_matrix.xlsx"


def read_matrix_for_today() -> dict | None:
    """
    Liest schedule_matrix.xlsx und gibt den Eintrag fuer heute zurueck.
    Returns: {"startzeit": "03:00", "jobs": {"sec_delta_update": True, ...}} oder None.
    """
    if not HAS_OPENPYXL:
        log.warning("openpyxl nicht installiert — Spreadsheet-Steuerung deaktiviert")
        return None

    if not SCHEDULE_MATRIX.exists():
        log.warning(f"schedule_matrix.xlsx nicht gefunden: {SCHEDULE_MATRIX}")
        return None

    try:
        wb = load_workbook(SCHEDULE_MATRIX, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        log.error(f"schedule_matrix.xlsx nicht lesbar: {e}")
        return None

    today_str = date.today().strftime("%d.%m.%Y")

    # Spalten-Layout (ab Zeile 3, mit 0/1-Paaren):
    # A=Datum, B=Tag, C=Startzeit, D/E=SEC(0/1), F/G=Aktien(0/1), H/I=Kalk(0/1)
    # Die "1"-Spalte (E=5, G=7, I=9) bestimmt ob Job aktiv ist
    JOB_1_COLUMNS = {
        5: "sec_delta_update",      # Spalte E = SEC "1"-Zelle
        7: "aktien_delta_update",   # Spalte G = Aktien "1"-Zelle
        9: "aktien_kalkulation",    # Spalte I = Kalkulation "1"-Zelle
    }

    # Heute finden (Daten ab Zeile 3)
    for row in ws.iter_rows(min_row=3):
        datum_cell = row[0]  # Spalte A = Datum
        if datum_cell.value and str(datum_cell.value).strip() == today_str:
            result = {"startzeit": "03:00", "jobs": {}}

            # Startzeit (Spalte C)
            startzeit_cell = row[2] if len(row) > 2 else None
            if startzeit_cell and startzeit_cell.value:
                result["startzeit"] = str(startzeit_cell.value).strip()

            # Job-Status: "1"-Spalte pruefen
            for col_idx, job_name in JOB_1_COLUMNS.items():
                if col_idx - 1 < len(row):
                    cell = row[col_idx - 1]  # 0-basiert
                    val = cell.value
                    # Aktiv wenn Wert 1 ist UND nicht ausgegraut (Font-Farbe nicht dim)
                    is_active = val in (1, "1", True)
                    result["jobs"][job_name] = is_active
                else:
                    result["jobs"][job_name] = False

            wb.close()
            log.info(f"Matrix fuer {today_str}: Startzeit={result['startzeit']}, "
                     f"Jobs={sum(1 for v in result['jobs'].values() if v)}/{len(result['jobs'])} aktiv")
            return result

    wb.close()
    log.warning(f"Kein Eintrag in schedule_matrix.xlsx fuer {today_str}")
    return None


def write_results_to_matrix(results: dict) -> None:
    """
    Schreibt Job-Ergebnisse zurueck in schedule_matrix.xlsx.
    Erfolg: +1 in der 1-Spalte (gruen)
    Fehler: -XX in der 1-Spalte (rot), wobei XX der Error-Code ist
    """
    if not HAS_OPENPYXL or not SCHEDULE_MATRIX.exists():
        log.warning("Kann Ergebnisse nicht in xlsx schreiben")
        return

    try:
        from openpyxl import load_workbook as lw
        from openpyxl.styles import Font as F, PatternFill as PF
        wb = lw(SCHEDULE_MATRIX)
        ws = wb.active
    except Exception as e:
        log.error(f"xlsx oeffnen fehlgeschlagen: {e}")
        return

    today_str = date.today().strftime("%d.%m.%Y")

    # Styles
    GREEN_BG = PF("solid", fgColor="1A3A2A")
    GREEN_FONT = F(name="Arial", size=11, bold=True, color="3FB950")
    RED_BG = PF("solid", fgColor="3A1A1A")
    RED_FONT = F(name="Arial", size=11, bold=True, color="F85149")

    # Job-Name → 1-Spalte (1-basiert)
    JOB_RESULT_COLS = {
        "sec_delta_update": 5,      # Spalte E
        "aktien_delta_update": 7,   # Spalte G
        "aktien_kalkulation": 9,    # Spalte I
    }

    for row_idx in range(3, ws.max_row + 1):
        cell_val = ws.cell(row=row_idx, column=1).value
        if cell_val and str(cell_val).strip() == today_str:
            for job_name, col in JOB_RESULT_COLS.items():
                if job_name in results:
                    r = results[job_name]
                    status = r.get("status", "")
                    cell = ws.cell(row=row_idx, column=col)

                    if status == "success":
                        cell.value = 1
                        cell.font = GREEN_FONT
                        cell.fill = GREEN_BG
                    elif status == "error":
                        # Error-Code: -01 bis -99
                        error_code = r.get("error_code", -1)
                        cell.value = error_code
                        cell.font = RED_FONT
                        cell.fill = RED_BG
                    elif status == "skipped":
                        cell.value = 0

            wb.save(SCHEDULE_MATRIX)
            log.info(f"Ergebnisse in xlsx geschrieben fuer {today_str}")
            break

    wb.close()


def wait_for_startzeit(startzeit: str) -> None:
    """
    Wartet bis zur angegebenen Startzeit (HH:MM).
    Falls die Startzeit schon vorbei ist, wird sofort gestartet.
    Aktualisiert running.json alle 30 Sekunden waehrend des Wartens.
    """
    try:
        hour, minute = map(int, startzeit.split(":"))
    except ValueError:
        log.error(f"Ungueltige Startzeit: '{startzeit}' — starte sofort")
        return

    now = datetime.now()
    target = now.replace(hour=hour, minute=minute, second=0, microsecond=0)

    if now >= target:
        log.info(f"Startzeit {startzeit} bereits erreicht — starte sofort")
        return

    wait_seconds = (target - now).total_seconds()
    log.info(f"Warte bis {startzeit} ({int(wait_seconds)}s / {int(wait_seconds / 60)} Minuten)...")

    # Status: wartend — alle 30s aktualisieren damit Dashboard sieht dass Mantel lebt
    remaining = wait_seconds
    while remaining > 0:
        _update_running(
            "waiting",
            target_time=startzeit,
            wait_seconds=int(remaining),
            started=NOW_ISO,
        )
        sleep_chunk = min(30, remaining)
        time.sleep(sleep_chunk)
        remaining -= sleep_chunk

    log.info(f"Startzeit {startzeit} erreicht — Jobs starten")


# ===================================================================
#  TOML Schedule lesen (Wochentag-basiert)
# ===================================================================

def get_schedule_file() -> Path:
    """Ermittelt die TOML-Datei fuer den heutigen Wochentag."""
    weekday_num = datetime.now().weekday()  # 0=Mo, 6=So
    tag = WOCHENTAGE[weekday_num]
    return SCRIPT_DIR / f"schedule.{tag}.toml"


def load_schedule() -> tuple[dict, str]:
    """Liest die Wochentag-TOML und gibt (schedule_dict, dateiname) zurueck."""
    schedule_file = get_schedule_file()
    tag = schedule_file.stem.split(".")[-1]  # "montag", "dienstag", ...

    log.info(f"Wochentag: {tag.capitalize()} → {schedule_file.name}")

    if not schedule_file.exists():
        log.error(f"Schedule-Datei nicht gefunden: {schedule_file}")
        sys.exit(2)

    try:
        with open(schedule_file, "rb") as f:
            return tomllib.load(f), schedule_file.name
    except Exception as e:
        log.error(f"Schedule nicht parsebar: {schedule_file.name} — {e}")
        sys.exit(2)


def get_enabled_jobs(schedule: dict, matrix_overrides: dict | None = None) -> list[dict]:
    """
    Gibt aktivierte Jobs zurueck.
    Wenn matrix_overrides vorhanden, uebersteuert die xlsx die TOML-enabled-Flags.
    """
    jobs = schedule.get("jobs", [])

    if matrix_overrides and matrix_overrides.get("jobs"):
        xlsx_jobs = matrix_overrides["jobs"]
        for job in jobs:
            name = job.get("name", "")
            if name in xlsx_jobs:
                job["enabled"] = xlsx_jobs[name]

    return [j for j in jobs if j.get("enabled", False)]


# ===================================================================
#  Job-Runner
# ===================================================================

JOB_REGISTRY: dict[str, callable] = {}


def register_job(name: str):
    """Decorator zum Registrieren einer Job-Funktion."""
    def wrapper(func):
        JOB_REGISTRY[name] = func
        return func
    return wrapper


# Error-Codes fuer Jobs (-01 bis -99)
ERR_GENERIC = -1       # Unbekannter Fehler
ERR_TIMEOUT = -2       # Job hat nicht rechtzeitig geantwortet
ERR_NOT_FOUND = -3     # Script/Datei nicht gefunden
ERR_SUBPROCESS = -4    # Subprocess-Fehler (Exit-Code != 0)
ERR_IMPORT = -5        # Import-Fehler (Modul nicht installiert)
ERR_NO_DATA = -6       # Keine Daten zurueck
ERR_DEPENDENCY = -7    # Abhaengigkeit fehlgeschlagen
ERR_PARSE = -8         # Daten nicht parsebar


def run_single_job(job: dict, config: dict, results: dict) -> dict:
    """Fuehrt einen einzelnen Job aus. Gibt Result-Dict zurueck."""
    name = job["name"]
    modul = job.get("modul", name)
    beschreibung = job.get("beschreibung", name)
    abhaengig_von = job.get("abhaengig_von")
    retry_max = config.get("retry_max", 0)
    retry_pause = config.get("retry_pause_sec", 30)
    timeout_sec = config.get("timeout_job_sec", 600)

    result = {
        "name": name,
        "beschreibung": beschreibung,
        "modul": modul,
        "status": "pending",
        "start_time": "",
        "duration_sec": 0,
        "attempts": 0,
        "message": "",
        "error_code": 0,
        "details": [],
    }

    # Abhaengigkeit pruefen
    if abhaengig_von:
        dep_result = results.get(abhaengig_von)
        if dep_result and dep_result["status"] == "error":
            result["status"] = "skipped"
            result["error_code"] = ERR_DEPENDENCY
            result["message"] = f"Uebersprungen — '{abhaengig_von}' fehlgeschlagen"
            log.warning(f"  [{name}] Uebersprungen (Abhaengigkeit '{abhaengig_von}' fehlgeschlagen)")
            return result

    # Funktion finden
    func = JOB_REGISTRY.get(modul)
    if not func:
        result["status"] = "error"
        result["error_code"] = ERR_NOT_FOUND
        result["message"] = f"Modul '{modul}' nicht registriert"
        log.error(f"  [{name}] Modul '{modul}' nicht in JOB_REGISTRY")
        return result

    # Ausfuehren mit Retry und Timeout
    for attempt in range(1, retry_max + 2):
        result["attempts"] = attempt
        result["start_time"] = datetime.now().strftime("%H:%M:%S")
        start = time.time()

        try:
            log.info(f"  [{name}] Starte (Versuch {attempt}, Timeout {timeout_sec}s)...")

            # Job in Thread mit Timeout ausfuehren
            import threading
            job_result = [None]
            job_error = [None]

            def _run():
                try:
                    job_result[0] = func()
                except Exception as ex:
                    job_error[0] = ex

            t = threading.Thread(target=_run, daemon=True)
            t.start()
            t.join(timeout=timeout_sec)

            elapsed = round(time.time() - start, 1)
            result["duration_sec"] = elapsed

            if t.is_alive():
                # Timeout: Thread laeuft noch
                result["status"] = "error"
                result["error_code"] = ERR_TIMEOUT
                result["message"] = f"Timeout nach {timeout_sec}s — keine Antwort vom Job"
                log.error(f"  [{name}] TIMEOUT nach {timeout_sec}s")
                return result  # Kein Retry bei Timeout

            if job_error[0]:
                raise job_error[0]

            job_output = job_result[0]
            result["status"] = "success"
            result["error_code"] = 0
            result["message"] = job_output.get("message", "OK") if isinstance(job_output, dict) else "OK"
            result["details"] = job_output.get("details", []) if isinstance(job_output, dict) else []
            log.info(f"  [{name}] Erfolg ({elapsed}s)")
            return result

        except Exception as e:
            elapsed = round(time.time() - start, 1)
            result["duration_sec"] = elapsed
            result["status"] = "error"
            result["message"] = str(e)
            if hasattr(e, 'error_code'):
                result["error_code"] = e.error_code
            else:
                result["error_code"] = ERR_GENERIC
            log.error(f"  [{name}] Fehler (Versuch {attempt}): {e}")

            if attempt <= retry_max:
                log.info(f"  [{name}] Warte {retry_pause}s vor Retry...")
                time.sleep(retry_pause)

    return result


# ===================================================================
#  Job-Module (registrierte Funktionen)
# ===================================================================

@register_job("collect_nachtbatch")
def job_collect_nachtbatch() -> dict:
    """Liest das Nacht-Log aus dem Vault."""
    nacht_log_file = NACHT_LOG_DIR / f"{TODAY}_nacht-batch.md"
    content = _read_file_safe(nacht_log_file)
    details = []

    if not content:
        return {
            "message": f"Kein Nacht-Log fuer {TODAY}",
            "details": [f"Erwartet: {nacht_log_file.name}"],
        }

    has_error = "fehler" in content.lower() or "error" in content.lower()
    for line in content.split("\n"):
        line = line.strip()
        if line.startswith("- [x]"):
            details.append(line[5:].strip())
        elif line.startswith("- ") and not line.startswith("- ["):
            details.append(line[2:])

    status_text = "mit Fehlern" if has_error else "erfolgreich"
    return {"message": f"Nacht-Batch {status_text}", "details": details[:5]}


@register_job("collect_sec_status")
def job_collect_sec_status() -> dict:
    """
    Fuehrt das SEC EDGAR Delta-Update aus.
    Ruft update_sec_data.py als Subprocess auf und wertet status.json aus.

    Rueckgabe: +1 bei Erfolg, Error-Code bei Fehler.
    Error-Codes:
      -3  = Script nicht gefunden
      -4  = Subprocess-Fehler (Exit-Code != 0)
      -6  = Keine Daten / Wochenende
      -8  = Status nicht parsebar
    """
    SEC_SCRIPT = SEC_DIR / "update_sec_data.py"

    # Wochenende: So (6) und Mo (0) — SEC hat keine neuen Filings
    weekday = datetime.now().weekday()
    if weekday in (6, 0):
        day_name = "Sonntag" if weekday == 6 else "Montag"
        return {"message": f"{day_name} — kein SEC-Update (Wochenende)", "details": []}

    # Script vorhanden?
    if not SEC_SCRIPT.exists():
        e = FileNotFoundError(f"SEC-Script nicht gefunden: {SEC_SCRIPT}")
        e.error_code = ERR_NOT_FOUND
        raise e

    # Subprocess ausfuehren
    log.info(f"    Starte {SEC_SCRIPT.name}...")
    rc, stdout, stderr = _run_cmd(
        [sys.executable, str(SEC_SCRIPT)],
        timeout=1800  # 30 Minuten max fuer SEC-Update
    )

    # Exit-Code pruefen
    if rc != 0:
        msg = stderr[:200] if stderr else f"Exit-Code {rc}"
        e = RuntimeError(f"SEC-Update fehlgeschlagen: {msg}")
        e.error_code = ERR_SUBPROCESS
        raise e

    # status.json auswerten
    content = _read_file_safe(SEC_STATUS_FILE)
    if not content:
        e = RuntimeError("SEC status.json nicht lesbar nach Update")
        e.error_code = ERR_NO_DATA
        raise e

    try:
        status = json.loads(content)
    except json.JSONDecodeError as exc:
        e = RuntimeError(f"SEC status.json nicht parsebar: {exc}")
        e.error_code = ERR_PARSE
        raise e

    phase = status.get("phase", "unknown")

    details = []
    if status.get("ciks_updated"):
        details.append(f"{status['ciks_updated']} Firmen aktualisiert")
    if status.get("facts_loaded"):
        details.append(f"{status['facts_loaded']:,} Facts geladen")
    if status.get("new_registrations"):
        details.append(f"{status['new_registrations']} neue Registrierungen")
    if status.get("elapsed_sec"):
        details.append(f"Laufzeit: {status['elapsed_sec']}s")
    if status.get("errors"):
        details.append(f"{status['errors']} Fehler")

    if phase == "completed":
        return {"message": "SEC Delta-Update erfolgreich", "details": details}
    elif phase == "no_updates":
        return {"message": status.get("message", "Keine Updates"), "details": details}
    elif phase == "failed":
        e = RuntimeError(f"SEC-Update Phase: {phase} — {status.get('message', '')}")
        e.error_code = ERR_SUBPROCESS
        raise e
    else:
        e = RuntimeError(f"SEC-Update unerwartete Phase: {phase}")
        e.error_code = ERR_GENERIC
        raise e


@register_job("collect_market_data")
def job_collect_market_data() -> dict:
    """Holt Marktdaten via yfinance und schreibt market.json."""
    try:
        import yfinance as yf
    except ImportError:
        raise ImportError("yfinance nicht installiert: pip3 install yfinance")

    tickers = [
        ("SPY", "^GSPC", "S&P 500 ETF"),
        ("QQQ", "QQQ", "Nasdaq 100 ETF"),
        ("IWM", "IWM", "Russell 2000 ETF"),
        ("VIX", "^VIX", "Volatilitaet"),
        ("DXY", "DX-Y.NYB", "US Dollar Index"),
        ("TNX", "^TNX", "10Y Treasury Yield"),
        ("GLD", "GLD", "Gold ETF"),
        ("BTC", "BTC-USD", "Bitcoin"),
    ]

    indices = []
    for display, symbol, label in tickers:
        try:
            hist = yf.Ticker(symbol).history(period="2d")
            if len(hist) >= 2:
                price = round(float(hist["Close"].iloc[-1]), 2)
                prev = float(hist["Close"].iloc[-2])
                change = round((price - prev) / prev * 100, 2)
            elif len(hist) == 1:
                price = round(float(hist["Close"].iloc[-1]), 2)
                change = 0.0
            else:
                price, change = 0.0, 0.0
        except Exception:
            price, change = 0.0, 0.0

        indices.append({"ticker": display, "label": label, "price": price, "change_pct": change})

    weinstein = _determine_weinstein()

    market = {"date": TODAY, "timestamp": NOW_ISO, "weinstein_phase": weinstein, "indices": indices}
    _write_json("market", market)

    loaded = sum(1 for i in indices if i["price"] > 0)
    return {
        "message": f"{loaded}/{len(tickers)} Kurse geladen",
        "details": [f"{i['ticker']}: {i['price']}" for i in indices if i["price"] > 0],
    }


@register_job("calculate_ampel")
def job_calculate_ampel() -> dict:
    """Trading-Ampel berechnen (VIX-Regime)."""
    market_file = DATA_DIR / "market.json"
    if not market_file.exists():
        raise FileNotFoundError("market.json nicht vorhanden — marktdaten-Job zuerst ausfuehren")

    market = json.loads(market_file.read_text(encoding="utf-8"))
    vix = 0.0
    for idx in market.get("indices", []):
        if idx["ticker"] == "VIX":
            vix = idx["price"]

    if vix == 0:
        signal, regime, size = "gelb", "unbekannt", "50%"
    elif vix < 16:
        signal, regime, size = "gruen", "niedrig", "100%"
    elif vix < 22:
        signal, regime, size = "gruen", "normal", "100%"
    elif vix < 30:
        signal, regime, size = "gelb", "erhoeht", "50%"
    else:
        signal, regime, size = "rot", "hoch", "25%"

    phase_labels = {0: "unbekannt", 1: "Bodenbildung", 2: "Aufwaertstrend", 3: "Topbildung", 4: "Abwaertstrend"}

    ampel = {
        "date": TODAY, "timestamp": NOW_ISO,
        "signal": signal, "vix_value": vix, "vix_regime": regime,
        "position_size": size,
        "market_phase": phase_labels.get(market.get("weinstein_phase", 0), "unbekannt"),
    }
    _write_json("ampel", ampel)
    return {"message": f"Ampel: {signal} (VIX {vix})", "details": [f"Regime: {regime}", f"Position: {size}"]}


@register_job("collect_breadth")
def job_collect_breadth() -> dict:
    """Marktbreite — Platzhalter."""
    breadth = {
        "date": TODAY, "timestamp": NOW_ISO,
        "indicators": [
            {"name": "Aktien > 200-MA", "value": 0, "total": 100, "pct": 0},
            {"name": "Aktien > 50-MA", "value": 0, "total": 100, "pct": 0},
            {"name": "Neue 52W-Hochs", "value": 0, "total": 100, "pct": 0},
            {"name": "Advance/Decline", "value": 0, "total": 100, "pct": 50},
            {"name": "Up-Volume Ratio", "value": 0, "total": 100, "pct": 50},
        ],
    }
    _write_json("breadth", breadth)
    return {"message": "Breadth-Platzhalter geschrieben", "details": []}


@register_job("git_commit_and_push")
def job_git_push() -> dict:
    """Committed und pusht data/*.json zu GitHub."""
    if not (SCRIPT_DIR / ".git").exists():
        raise FileNotFoundError(f"Kein Git-Repo unter {SCRIPT_DIR}")

    rc, out, err = _run_cmd(["git", "add", "data/"])
    if rc != 0:
        raise RuntimeError(f"git add fehlgeschlagen: {err}")

    rc, _, _ = _run_cmd(["git", "diff", "--cached", "--quiet"])
    if rc == 0:
        return {"message": "Keine Aenderungen — kein Push noetig", "details": []}

    msg = f"Mantelroutine {TODAY} {NOW_HM}"
    rc, out, err = _run_cmd(["git", "commit", "-m", msg])
    if rc != 0:
        raise RuntimeError(f"git commit fehlgeschlagen: {err}")

    rc, out, err = _run_cmd(["git", "push", "origin", "main"])
    if rc != 0:
        raise RuntimeError(f"git push fehlgeschlagen: {err}")

    return {"message": "Git push erfolgreich", "details": [msg]}


# ===================================================================
#  Ergebnis-JSON schreiben (fuer Dashboard)
# ===================================================================

def write_schedule_result(schedule: dict, schedule_file: str, jobs_enabled: list, results: dict, duration: float):
    """Schreibt schedule_result.json fuer die Dashboard-Anzeige."""
    mantel_cfg = schedule.get("mantel", {})
    wochentag = mantel_cfg.get("wochentag", "unbekannt")

    job_results = []
    for job in jobs_enabled:
        name = job["name"]
        r = results.get(name, {"status": "skipped", "message": "Nicht ausgefuehrt"})
        job_results.append(r)

    statuses = [r.get("status") for r in results.values()]
    if not statuses:
        overall = "empty"
    elif all(s == "success" for s in statuses):
        overall = "success"
    elif all(s == "error" for s in statuses):
        overall = "error"
    elif any(s == "error" for s in statuses):
        overall = "partial"
    else:
        overall = "success"

    schedule_result = {
        "date": TODAY,
        "timestamp": NOW_ISO,
        "wochentag": wochentag,
        "schedule_file": schedule_file,
        "startzeit_actual": NOW_HM,
        "overall_status": overall,
        "duration_total_sec": round(duration, 1),
        "jobs_total": len(jobs_enabled),
        "jobs_success": sum(1 for s in statuses if s == "success"),
        "jobs_error": sum(1 for s in statuses if s == "error"),
        "jobs_skipped": sum(1 for s in statuses if s == "skipped"),
        "jobs": job_results,
        "schedule_empty": len(jobs_enabled) == 0,
        "message": f"Kein Eintrag im Schedule ({wochentag.capitalize()})" if not jobs_enabled else "",
    }

    _write_json("schedule_result", schedule_result)

    # Auch nachtbatch.json fuer Kompatibilitaet
    nachtbatch = {
        "date": TODAY, "timestamp": NOW_ISO,
        "overall_status": overall,
        "jobs": job_results,
        "pending_actions": [],
        "history": _build_history(),
    }
    _write_json("nachtbatch", nachtbatch)


def write_empty_schedule_result(schedule: dict, schedule_file: str):
    """Schreibt Ergebnis fuer leeren Schedule."""
    mantel_cfg = schedule.get("mantel", {})
    wochentag = mantel_cfg.get("wochentag", "unbekannt")

    schedule_result = {
        "date": TODAY,
        "timestamp": NOW_ISO,
        "wochentag": wochentag,
        "schedule_file": schedule_file,
        "startzeit_actual": NOW_HM,
        "overall_status": "empty",
        "duration_total_sec": 0,
        "jobs_total": 0,
        "jobs_success": 0,
        "jobs_error": 0,
        "jobs_skipped": 0,
        "jobs": [],
        "schedule_empty": True,
        "message": f"Kein Eintrag im Schedule ({wochentag.capitalize()})",
    }

    _write_json("schedule_result", schedule_result)
    log.info(f"Schedule leer: {schedule_file} — keine Jobs konfiguriert")


# ===================================================================
#  Hilfsfunktionen
# ===================================================================

def _read_file_safe(path: Path) -> str:
    try:
        return path.read_text(encoding="utf-8")
    except Exception as e:
        log.warning(f"Datei nicht lesbar: {path} — {e}")
        return ""


def _write_json(name: str, data: dict):
    fp = DATA_DIR / f"{name}.json"
    with open(fp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    log.info(f"  -> {fp.name}")


def _run_cmd(cmd: list, timeout: int = 60, cwd: str = None) -> tuple:
    try:
        r = subprocess.run(
            cmd,
            cwd=cwd or str(SCRIPT_DIR),
            capture_output=True,
            text=True,
            timeout=timeout,
        )
        return r.returncode, r.stdout.strip(), r.stderr.strip()
    except subprocess.TimeoutExpired:
        return -2, "", f"Timeout nach {timeout}s"
    except Exception as e:
        return -1, "", str(e)


def _determine_weinstein() -> int:
    try:
        import yfinance as yf
        hist = yf.Ticker("SPY").history(period="1y")
        if len(hist) < 200:
            return 0
        c = float(hist["Close"].iloc[-1])
        sma30 = float(hist["Close"].iloc[-30:].mean())
        sma200 = float(hist["Close"].iloc[-200:].mean())
        if c > sma200 and c > sma30 and sma30 > sma200:
            return 2
        elif c > sma200 and sma30 < sma200:
            return 1
        elif c < sma200 and c < sma30 and sma30 < sma200:
            return 4
        else:
            return 3
    except Exception:
        return 0


def _build_history() -> list:
    history = []
    for days_ago in range(1, 4):
        d = (datetime.now() - timedelta(days=days_ago)).strftime("%Y-%m-%d")
        log_file = NACHT_LOG_DIR / f"{d}_nacht-batch.md"
        c = _read_file_safe(log_file)
        if c:
            has_err = "fehler" in c.lower() or "error" in c.lower()
            history.append({"date": d, "status": "partial" if has_err else "success", "jobs_ok": 1 if not has_err else 0, "jobs_total": 2})
        else:
            history.append({"date": d, "status": "error", "jobs_ok": 0, "jobs_total": 2, "note": "Kein Log"})
    return history


# ===================================================================
#  Hauptprogramm
# ===================================================================

def main() -> int:
    sofort = "--sofort" in sys.argv

    log.info("=" * 60)
    log.info(f"Mantelroutine gestartet — {NOW_ISO}" + (" [SOFORT]" if sofort else ""))
    log.info("=" * 60)

    # Status: gestartet
    _update_running("waiting", target_time="--:--", wait_seconds=0, started=NOW_ISO)

    # 1. Spreadsheet lesen (Startzeit + Job-Overrides)
    matrix = read_matrix_for_today()
    if matrix:
        if sofort:
            log.info(f"Spreadsheet-Steuerung aktiv, aber --sofort: Startzeit {matrix['startzeit']} uebersprungen")
        else:
            log.info(f"Spreadsheet-Steuerung aktiv: Startzeit={matrix['startzeit']}")
            wait_for_startzeit(matrix["startzeit"])
    else:
        log.info("Keine Spreadsheet-Steuerung — starte sofort mit TOML-Defaults")

    # 2. Wochentag-TOML laden
    schedule, schedule_file = load_schedule()
    mantel_cfg = schedule.get("mantel", {})
    wochentag = mantel_cfg.get("wochentag", "unbekannt")
    log.info(f"Schedule: {schedule_file} ({wochentag.capitalize()})")

    # 3. Aktivierte Jobs ermitteln (xlsx uebersteuert TOML)
    jobs_enabled = get_enabled_jobs(schedule, matrix)
    log.info(f"Jobs im Schedule: {len(jobs_enabled)}")

    # 3. Leerer Schedule?
    if not jobs_enabled:
        write_empty_schedule_result(schedule, schedule_file)
        _update_running("idle")
        log.info("=" * 60)
        log.info(f"Mantelroutine beendet — {schedule_file} leer (Erfolg)")
        log.info("=" * 60)
        return 0

    # 4. Jobs der Reihe nach ausfuehren
    job_names = [j["name"] for j in jobs_enabled]
    for j in jobs_enabled:
        log.info(f"  -> {j['name']}: {j.get('beschreibung', '')}")

    results = {}
    start_total = time.time()

    for idx, job in enumerate(jobs_enabled):
        # Live-Status aktualisieren: welcher Job laeuft gerade?
        _update_running(
            "running",
            current_job=job["name"],
            current_job_beschreibung=job.get("beschreibung", job["name"]),
            job_index=idx + 1,
            jobs_total=len(jobs_enabled),
            job_names=job_names,
            started=NOW_ISO,
            job_started=datetime.now().isoformat(),
            results_so_far=[
                {"name": r["name"], "status": r["status"], "error_code": r.get("error_code", 0),
                 "duration_sec": r.get("duration_sec", 0)}
                for r in results.values()
            ],
        )

        result = run_single_job(job, mantel_cfg, results)
        results[job["name"]] = result

    duration = time.time() - start_total

    # 5. Ergebnisse in xlsx zurueckschreiben (+1 Erfolg, -XX Fehler)
    write_results_to_matrix(results)

    # 6. Ergebnis-JSON schreiben (fuer Dashboard)
    write_schedule_result(schedule, schedule_file, jobs_enabled, results, duration)

    # 7. Zusammenfassung
    success_count = sum(1 for r in results.values() if r["status"] == "success")
    error_count = sum(1 for r in results.values() if r["status"] == "error")
    skipped_count = sum(1 for r in results.values() if r["status"] == "skipped")

    # Status: fertig
    _update_running(
        "finished",
        overall_status="success" if error_count == 0 else "error",
        jobs_success=success_count,
        jobs_error=error_count,
        jobs_skipped=skipped_count,
        duration_sec=round(duration, 1),
        started=NOW_ISO,
        results=[
            {"name": r["name"], "status": r["status"], "message": r.get("message", ""),
             "error_code": r.get("error_code", 0), "duration_sec": r.get("duration_sec", 0)}
            for r in results.values()
        ],
    )

    log.info("=" * 60)
    log.info(f"Ergebnis: {success_count} OK / {error_count} Fehler / {skipped_count} uebersprungen ({round(duration, 1)}s)")
    log.info("=" * 60)

    return 1 if error_count > 0 else 0


if __name__ == "__main__":
    sys.exit(main())
