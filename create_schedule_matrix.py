#!/usr/bin/env python3
"""
Erstellt schedule_matrix.xlsx — die Kalender-Steuerungstabelle.
Laufender + naechster Monat, Mo-So.

Zellenwerte pro Job (Spalten-Paar [0] [1]):
  leer/0 = nicht geplant (grau)
  1      = geplant, noch nicht gelaufen (gruen)
  +1     = erfolgreich gelaufen (gruen, bleibt)
  -01..-99 = Fehler mit Error-Code (rot)

Vorbelegung:
  Mo + So = keine Laeufe (0)
  Di - Sa = alle 3 Jobs aktiv (1), Startzeit 03:00
  Vergangene Tage = 0

Aufruf: python3 create_schedule_matrix.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
from pathlib import Path
import calendar

SCRIPT_DIR = Path(__file__).parent.resolve()
OUTPUT = SCRIPT_DIR / "schedule_matrix.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "Schedule"

# --- Styles ---
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(name="Arial", bold=True, color="E6EDF3", size=10)
SUBHEADER_FONT = Font(name="Arial", bold=False, color="8B949E", size=9)
TIME_FONT = Font(name="Arial", size=10, color="58A6FF")

# Job-Zellen: aktiv vs inaktiv
GREEN_BG = PatternFill("solid", fgColor="1A3A2A")    # sanftes Gruen
GREEN_FONT = Font(name="Arial", size=11, bold=True, color="3FB950")
RED_BG = PatternFill("solid", fgColor="3A1A1A")      # sanftes Rot
RED_FONT = Font(name="Arial", size=11, bold=True, color="F85149")
DIM_BG = PatternFill("solid", fgColor="0D1117")      # inaktiv/ausgegraut
DIM_FONT = Font(name="Arial", size=10, color="30363D")

WEEKEND_FILL = PatternFill("solid", fgColor="161B22")
NORMAL_FILL = PatternFill("solid", fgColor="0D1117")
TODAY_FILL = PatternFill("solid", fgColor="1C2333")

BORDER = Border(
    left=Side(style="thin", color="30363D"),
    right=Side(style="thin", color="30363D"),
    top=Side(style="thin", color="30363D"),
    bottom=Side(style="thin", color="30363D"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

# --- Spalten-Layout ---
# A=Datum, B=Tag, C=Startzeit, D/E=SEC, F/G=Aktien Delta, H/I=Aktien Kalk
# D=0-Zelle, E=1-Zelle (fuer jeden Job)

# Header Zeile 1: Hauptheader
headers_row1 = [
    ("Datum", 14), ("Tag", 12), ("Startzeit", 12),
    ("SEC Delta Update", 8), ("", 8),
    ("Aktien Delta Update", 8), ("", 8),
    ("Aktien Kalkulation", 8), ("", 8),
]

for col_idx, (header, width) in enumerate(headers_row1, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# Merged headers fuer Job-Paare
ws.merge_cells("D1:E1")
ws.merge_cells("F1:G1")
ws.merge_cells("H1:I1")

# Header Zeile 2: Sub-Header (0/1 Labels)
subheaders = ["", "", "", "0", "1", "0", "1", "0", "1"]
for col_idx, sh in enumerate(subheaders, 1):
    cell = ws.cell(row=2, column=col_idx, value=sh)
    cell.font = SUBHEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER

# --- Kalender ---
today = date.today()
year, month = today.year, today.month
first_of_month = date(year, month, 1)
start = first_of_month - timedelta(days=first_of_month.weekday())

if month == 12:
    next_month_last = date(year + 1, 1, calendar.monthrange(year + 1, 1)[1])
else:
    next_month_last = date(year, month + 1, calendar.monthrange(year, month + 1)[1])
end = next_month_last + timedelta(days=(6 - next_month_last.weekday()) % 7)

WOCHENTAGE = ["Montag", "Dienstag", "Mittwoch", "Donnerstag", "Freitag", "Samstag", "Sonntag"]

row = 3  # Daten ab Zeile 3
current = start
while current <= end:
    wt = current.weekday()
    is_weekend = wt >= 5
    is_today = current == today
    is_past = current < today
    fill = TODAY_FILL if is_today else (WEEKEND_FILL if is_weekend else NORMAL_FILL)

    # A: Datum
    c = ws.cell(row=row, column=1, value=current.strftime("%d.%m.%Y"))
    c.font = Font(name="Arial", size=10, color="E6EDF3", bold=is_today)
    c.fill = fill; c.border = BORDER; c.alignment = LEFT

    # B: Wochentag
    c = ws.cell(row=row, column=2, value=WOCHENTAGE[wt])
    c.font = Font(name="Arial", size=10, color="58A6FF" if is_today else ("6E7681" if is_weekend else "8B949E"))
    c.fill = fill; c.border = BORDER; c.alignment = LEFT

    # C: Startzeit
    c = ws.cell(row=row, column=3, value="03:00")
    c.font = TIME_FONT; c.fill = fill; c.border = BORDER; c.alignment = CENTER

    # Job-Paare: D/E, F/G, H/I
    # Aktiver Wert hat Zahl + Farbe, inaktiver ist leer + dim
    # Vergangene Tage: AUS (0 in D/F/H rot, E/G/I leer)
    # Zukunft/Heute: AN (D/F/H leer, 1 in E/G/I gruen)
    for pair_start in [4, 6, 8]:
        col_0 = pair_start      # 0-Zelle
        col_1 = pair_start + 1  # 1-Zelle

        # Mo (0) und So (6) = keine Laeufe, sonst aktiv
        job_active = not is_past and wt not in (0, 6)

        if job_active:
            # AN: 0-Zelle leer (dim), 1-Zelle hat Wert 1 (gruen)
            c0 = ws.cell(row=row, column=col_0)
            c0.font = DIM_FONT; c0.fill = DIM_BG; c0.border = BORDER; c0.alignment = CENTER
            c1 = ws.cell(row=row, column=col_1, value=1)
            c1.font = GREEN_FONT; c1.fill = GREEN_BG; c1.border = BORDER; c1.alignment = CENTER
        else:
            # AUS: 0-Zelle hat Wert 0 (rot bei Vergangenheit, dim bei So/Mo)
            c0 = ws.cell(row=row, column=col_0, value=0)
            if is_past:
                c0.font = RED_FONT; c0.fill = RED_BG
            else:
                c0.font = DIM_FONT; c0.fill = DIM_BG
            c0.border = BORDER; c0.alignment = CENTER
            c1 = ws.cell(row=row, column=col_1)
            c1.font = DIM_FONT; c1.fill = DIM_BG; c1.border = BORDER; c1.alignment = CENTER

    row += 1
    current += timedelta(days=1)

# Freeze: Header + Sub-Header
ws.freeze_panes = "A3"
ws.sheet_properties.tabColor = "58A6FF"

wb.save(OUTPUT)
print(f"Erstellt: {OUTPUT} ({row - 3} Kalendertage, {start} bis {end})")
