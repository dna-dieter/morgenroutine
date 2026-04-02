#!/usr/bin/env python3
"""
Morgenroutine — Lokaler Dashboard-Server
==========================================
Startet einen Mini-Webserver auf localhost:8787.
Zeigt die schedule_matrix.xlsx als Kalender mit Startzeiten und Jobs.
Erlaubt Bearbeitung direkt im Browser.

Aufruf:  python3 dashboard_server.py
Oeffnet: http://localhost:8787

Beenden: Ctrl+C oder "Schliessen und Stoppen"-Button
"""

import http.server
import json
import os
import signal
import sys
import threading
import tomllib
import webbrowser
from datetime import datetime, date, timedelta
from pathlib import Path
from http import HTTPStatus

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("WARNUNG: openpyxl nicht installiert — pip3 install openpyxl")

PORT = 8787
SCRIPT_DIR = Path(__file__).parent.resolve()
SCHEDULE_MATRIX = SCRIPT_DIR / "schedule_matrix.xlsx"
RUNNING_FILE = SCRIPT_DIR / "data" / "running.json"

WOCHENTAGE = ["Montag", "Dienstag", "Mittwoch", "Donnerstag", "Freitag", "Samstag", "Sonntag"]

# Spalten-Layout: "1"-Spalte bestimmt ob Job aktiv (0-basierter Index)
# D/E=SEC(0/1), F/G=Aktien(0/1), H/I=Kalk(0/1)
# Die "1"-Spalten sind E=4, G=6, I=8 (0-basiert)
JOB_1_COLS = {
    4: "sec_delta_update",
    6: "aktien_delta_update",
    8: "aktien_kalkulation",
}


def read_xlsx_schedule() -> dict:
    """Liest die gesamte schedule_matrix.xlsx und gibt strukturierte Daten zurueck."""
    if not HAS_OPENPYXL or not SCHEDULE_MATRIX.exists():
        return {"error": "schedule_matrix.xlsx nicht gefunden", "weeks": [], "today": date.today().isoformat()}

    wb = load_workbook(SCHEDULE_MATRIX, data_only=True)
    ws = wb.active

    # Header lesen
    headers = {}
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value:
            headers[col_idx] = str(cell.value).strip()

    # Alle Tage lesen (Daten ab Zeile 3, Zeile 1=Header, Zeile 2=Sub-Header)
    days = []
    today = date.today()

    for row in ws.iter_rows(min_row=3):
        datum_val = row[0].value
        if not datum_val:
            continue

        datum_str = str(datum_val).strip()
        try:
            d = datetime.strptime(datum_str, "%d.%m.%Y").date()
        except ValueError:
            continue

        tag_val = row[1].value if len(row) > 1 else ""
        startzeit = str(row[2].value).strip() if len(row) > 2 and row[2].value else "03:00"

        jobs = {}
        for col_idx, job_key in JOB_1_COLS.items():
            if col_idx < len(row) and row[col_idx].value is not None:
                val = row[col_idx].value
                try:
                    num = int(val) if not isinstance(val, (int, float)) else val
                except (ValueError, TypeError):
                    num = 0
                # 1 oder positiv = aktiv/erfolg, negativ = fehler, 0/leer = inaktiv
                jobs[job_key] = {"value": int(num), "active": num > 0, "error": num < 0}
            else:
                jobs[job_key] = {"value": 0, "active": False, "error": False}

        days.append({
            "datum": datum_str,
            "datum_iso": d.isoformat(),
            "wochentag": str(tag_val).strip() if tag_val else WOCHENTAGE[d.weekday()],
            "wochentag_kurz": WOCHENTAGE[d.weekday()][:2],
            "startzeit": startzeit,
            "jobs": jobs,
            "is_today": d == today,
            "is_past": d < today,
            "is_weekend": d.weekday() >= 5,
        })

    wb.close()

    # In Wochen gruppieren (Mo-So)
    weeks = []
    current_week = None

    for day in days:
        d = datetime.strptime(day["datum"], "%d.%m.%Y").date()
        kw = d.isocalendar()[1]
        kw_label = f"KW {kw}"

        if not current_week or current_week["label"] != kw_label:
            if current_week:
                weeks.append(current_week)
            current_week = {"label": kw_label, "days": []}

        current_week["days"].append(day)

    if current_week:
        weeks.append(current_week)

    # Letztes Ergebnis lesen
    result_file = SCRIPT_DIR / "data" / "schedule_result.json"
    last_result = {}
    if result_file.exists():
        try:
            last_result = json.loads(result_file.read_text(encoding="utf-8"))
        except Exception:
            pass

    return {"weeks": weeks, "today": today.isoformat(), "last_result": last_result}


def read_running_status() -> dict:
    """Liest running.json und gibt den aktuellen Mantel-Status zurueck."""
    if not RUNNING_FILE.exists():
        return {"state": "idle"}

    try:
        content = RUNNING_FILE.read_text(encoding="utf-8")
        data = json.loads(content)

        # Pruefen ob running.json noch aktuell ist (nicht aelter als 5 Minuten)
        updated = data.get("updated", "")
        if updated:
            try:
                last = datetime.fromisoformat(updated)
                age_sec = (datetime.now() - last).total_seconds()
                data["age_sec"] = round(age_sec, 1)
                # Wenn Prozess laenger als 5 Min nicht geschrieben hat → vermutlich tot
                if age_sec > 300 and data.get("state") in ("waiting", "running"):
                    data["stale"] = True
            except ValueError:
                pass

        # Pruefen ob PID noch lebt
        pid = data.get("pid")
        if pid:
            try:
                os.kill(pid, 0)  # Signal 0 = nur pruefen ob Prozess existiert
                data["pid_alive"] = True
            except (OSError, ProcessLookupError):
                data["pid_alive"] = False
                if data.get("state") in ("waiting", "running"):
                    data["state"] = "crashed"

        return data
    except (json.JSONDecodeError, Exception):
        return {"state": "idle"}


# Styles fuer xlsx-Schreibvorgaenge
try:
    from openpyxl.styles import Font as XlFont, PatternFill as XlFill
    GREEN_BG = XlFill("solid", fgColor="1A3A2A")
    GREEN_FONT = XlFont(name="Arial", size=11, bold=True, color="3FB950")
    RED_BG = XlFill("solid", fgColor="3A1A1A")
    RED_FONT = XlFont(name="Arial", size=11, bold=True, color="F85149")
    DIM_BG = XlFill("solid", fgColor="0D1117")
    DIM_FONT = XlFont(name="Arial", size=10, color="30363D")
except ImportError:
    pass

# Job-Key → (0-Spalte 1-basiert, 1-Spalte 1-basiert)
JOB_COL_PAIRS = {
    "sec_delta_update": (4, 5),
    "aktien_delta_update": (6, 7),
    "aktien_kalkulation": (8, 9),
}


def save_xlsx_day(datum_str: str, startzeit: str, jobs: dict) -> bool:
    """Aktualisiert einen einzelnen Tag in der schedule_matrix.xlsx."""
    if not HAS_OPENPYXL or not SCHEDULE_MATRIX.exists():
        return False

    wb = load_workbook(SCHEDULE_MATRIX)
    ws = wb.active

    # Zeile fuer diesen Tag finden (Daten ab Zeile 3)
    for row_idx in range(3, ws.max_row + 1):
        cell_val = ws.cell(row=row_idx, column=1).value
        if cell_val and str(cell_val).strip() == datum_str:
            # Startzeit schreiben (Spalte C)
            ws.cell(row=row_idx, column=3, value=startzeit)

            # Jobs schreiben: aktiver Wert hat Zahl+Farbe, inaktiver ist leer+dim
            for job_key, (col_0, col_1) in JOB_COL_PAIRS.items():
                if job_key in jobs:
                    val = jobs[job_key]
                    # val kann bool (toggle) oder int (error-code) sein
                    if isinstance(val, bool):
                        is_on = val
                    elif isinstance(val, (int, float)):
                        is_on = val > 0
                    else:
                        is_on = bool(val)
                    c0 = ws.cell(row=row_idx, column=col_0)
                    c1 = ws.cell(row=row_idx, column=col_1)
                    if is_on:
                        c0.value = None; c0.font = DIM_FONT; c0.fill = DIM_BG
                        c1.value = 1; c1.font = GREEN_FONT; c1.fill = GREEN_BG
                    else:
                        c0.value = 0; c0.font = RED_FONT; c0.fill = RED_BG
                        c1.value = None; c1.font = DIM_FONT; c1.fill = DIM_BG

            wb.save(SCHEDULE_MATRIX)
            wb.close()
            return True

    wb.close()
    return False


class DashboardHandler(http.server.BaseHTTPRequestHandler):
    def log_message(self, format, *args):
        pass

    def do_GET(self):
        if self.path == "/" or self.path == "/index.html":
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.end_headers()
            self.wfile.write(HTML_PAGE.encode("utf-8"))
        elif self.path == "/api/schedule":
            data = read_xlsx_schedule()
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "application/json")
            self.end_headers()
            self.wfile.write(json.dumps(data, ensure_ascii=False).encode("utf-8"))
        elif self.path == "/api/status":
            data = read_running_status()
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "application/json")
            self.send_header("Cache-Control", "no-cache")
            self.end_headers()
            self.wfile.write(json.dumps(data, ensure_ascii=False).encode("utf-8"))
        else:
            self.send_error(HTTPStatus.NOT_FOUND)

    def do_POST(self):
        if self.path == "/api/shutdown":
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "application/json")
            self.end_headers()
            self.wfile.write(json.dumps({"ok": True}).encode("utf-8"))
            threading.Thread(target=lambda: (self.server.shutdown(), os._exit(0)), daemon=True).start()
            return
        elif self.path == "/api/save":
            length = int(self.headers.get("Content-Length", 0))
            body = json.loads(self.rfile.read(length))
            datum = body.get("datum", "")
            startzeit = body.get("startzeit", "03:00")
            jobs = body.get("jobs", {})
            ok = save_xlsx_day(datum, startzeit, jobs)
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "application/json")
            self.end_headers()
            self.wfile.write(json.dumps({"ok": ok, "datum": datum}).encode("utf-8"))
        elif self.path == "/api/trigger":
            # Alte running.json loeschen damit Polling nicht den alten State liest
            try:
                RUNNING_FILE.unlink(missing_ok=True)
            except Exception:
                pass
            # Mantel sofort starten (im Hintergrund, ohne auf Startzeit zu warten)
            import subprocess
            mantel = SCRIPT_DIR / "mantel.py"
            subprocess.Popen([sys.executable, str(mantel), "--sofort"], cwd=str(SCRIPT_DIR))
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "application/json")
            self.end_headers()
            self.wfile.write(json.dumps({"ok": True, "message": "Mantel gestartet"}).encode("utf-8"))
        else:
            self.send_error(HTTPStatus.NOT_FOUND)


HTML_PAGE = r"""<!DOCTYPE html>
<html lang="de">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Mantelroutine — Schedule Dashboard</title>
<style>
:root{--bg:#0d1117;--card:#161b22;--card-hover:#1c2333;--border:#30363d;--text:#e6edf3;--text2:#8b949e;--muted:#6e7681;--green:#3fb950;--red:#f85149;--yellow:#d29922;--blue:#58a6ff;--purple:#bc8cff}
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Helvetica,sans-serif;background:var(--bg);color:var(--text);padding:20px}

.top-bar{display:flex;justify-content:space-between;align-items:center;margin-bottom:20px}
h1{font-size:1.4rem} h1 span{color:var(--blue)}
.top-buttons{display:flex;gap:8px}

/* Live-Status-Banner */
.live-status{background:var(--card);border:1px solid var(--border);border-radius:8px;padding:12px 16px;margin-bottom:10px;display:none;align-items:center;gap:12px;font-size:.85rem}
.live-status.visible{display:flex}
.live-status.state-running{border-color:var(--blue);background:linear-gradient(90deg,rgba(88,166,255,.05),rgba(88,166,255,.12))}
.live-status.state-waiting{border-color:var(--yellow);background:linear-gradient(90deg,rgba(210,153,34,.05),rgba(210,153,34,.12))}
.live-status.state-finished{border-color:var(--green);background:linear-gradient(90deg,rgba(63,185,80,.05),rgba(63,185,80,.12))}
.live-status.state-crashed{border-color:var(--red);background:linear-gradient(90deg,rgba(248,81,73,.05),rgba(248,81,73,.12))}
.live-status.state-starting{border-color:var(--purple);background:linear-gradient(90deg,rgba(188,140,255,.05),rgba(188,140,255,.12))}

.pulse{width:12px;height:12px;border-radius:50%;flex-shrink:0;animation:pulse 1.5s ease-in-out infinite}
.pulse.blue{background:var(--blue);box-shadow:0 0 8px rgba(88,166,255,.5)}
.pulse.yellow{background:var(--yellow);box-shadow:0 0 8px rgba(210,153,34,.5)}
.pulse.green{background:var(--green);box-shadow:0 0 8px rgba(63,185,80,.5);animation:none}
.pulse.red{background:var(--red);box-shadow:0 0 8px rgba(248,81,73,.5);animation:none}
@keyframes pulse{0%,100%{opacity:1;transform:scale(1)}50%{opacity:.5;transform:scale(.8)}}

.status-text{flex:1}
.status-label{font-weight:700;margin-bottom:2px}
.status-detail{color:var(--text2);font-size:.8rem}
.status-timer{display:flex;flex-direction:column;align-items:flex-end;gap:2px;flex-shrink:0;margin-left:8px}
.timer-value{font-size:1.1rem;font-weight:700;font-family:monospace;color:var(--blue)}
.timer-label{font-size:.65rem;color:var(--text2);text-transform:uppercase;letter-spacing:.5px}

.job-chips{display:flex;gap:4px;margin-top:4px}
.job-chip{font-size:.7rem;padding:2px 8px;border-radius:10px;border:1px solid var(--border)}
.job-chip.done{background:rgba(63,185,80,.15);border-color:var(--green);color:var(--green)}
.job-chip.active{background:rgba(88,166,255,.15);border-color:var(--blue);color:var(--blue);animation:pulse 1.5s ease-in-out infinite}
.job-chip.pending{color:var(--muted)}
.job-chip.error{background:rgba(248,81,73,.15);border-color:var(--red);color:var(--red)}

.last-run{background:var(--card);border:1px solid var(--border);border-radius:8px;padding:12px 16px;margin-bottom:20px;font-size:.85rem;display:flex;align-items:center;gap:10px}
.dot{width:12px;height:12px;border-radius:50%;display:inline-block;flex-shrink:0}
.dot.empty,.dot.error{background:var(--red);box-shadow:0 0 8px rgba(248,81,73,.4)}
.dot.success{background:var(--green);box-shadow:0 0 8px rgba(63,185,80,.4)}
.dot.partial{background:var(--yellow);box-shadow:0 0 8px rgba(210,153,34,.4)}

.kw-label{font-size:.9rem;font-weight:700;color:var(--blue);margin:16px 0 8px;padding:4px 0;border-bottom:1px solid var(--border)}
.week-grid{display:grid;grid-template-columns:repeat(7,1fr);gap:6px;margin-bottom:12px}

.day-card{background:var(--card);border:1px solid var(--border);border-radius:8px;padding:10px;min-height:130px;cursor:pointer;transition:border-color .2s}
.day-card:hover{border-color:var(--blue)}
.day-card.today{border-color:var(--blue);border-width:2px}
.day-card.past{opacity:.5}
.day-card.weekend{background:#12161e}

.day-name{font-size:.7rem;color:var(--text2);text-transform:uppercase;letter-spacing:.5px}
.day-date{font-size:1rem;font-weight:700;margin:2px 0 4px}
.day-time{font-size:.85rem;color:var(--blue);margin-bottom:6px}

.job-row-mini{display:flex;align-items:center;gap:4px;font-size:.7rem;margin-bottom:2px}
.job-dot{width:8px;height:8px;border-radius:50%;flex-shrink:0}
.job-dot.active{background:var(--green)}
.job-dot.inactive{background:var(--muted)}
.job-dot.error{background:var(--red);box-shadow:0 0 6px rgba(248,81,73,.3)}
.job-error-code{font-size:.6rem;color:var(--red);font-weight:700;margin-left:2px}
.job-label{color:var(--text2)}
.job-label.active{color:var(--green)}
.job-label.error{color:var(--red)}
.job-toggle-click{cursor:pointer;padding:2px 4px;border-radius:4px;transition:background .15s}
.job-toggle-click:hover{background:rgba(88,166,255,.1)}

/* Modal */
.modal-overlay{display:none;position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(0,0,0,.7);z-index:100;justify-content:center;align-items:center}
.modal-overlay.open{display:flex}
.modal{background:var(--card);border:1px solid var(--border);border-radius:12px;padding:24px;width:500px;max-height:80vh;overflow-y:auto}
.modal h2{font-size:1.1rem;margin-bottom:16px} .modal h2 span{color:var(--blue)}

.form-group{margin-bottom:14px}
.form-group label{display:block;font-size:.8rem;color:var(--text2);margin-bottom:4px}
.form-group input[type=text]{background:var(--bg);border:1px solid var(--border);border-radius:6px;color:var(--blue);padding:8px 12px;font-size:1rem;width:100px;text-align:center;font-family:monospace}

.job-toggle{display:flex;align-items:center;gap:10px;padding:8px 12px;background:var(--bg);border-radius:6px;margin-bottom:6px}
.job-toggle input[type=checkbox]{accent-color:var(--green);width:18px;height:18px}
.job-toggle .job-name{font-size:.9rem;flex:1}
.job-toggle .job-desc{font-size:.75rem;color:var(--muted)}

.btn{padding:6px 16px;border:1px solid var(--border);border-radius:6px;background:var(--card);color:var(--text);cursor:pointer;font-size:.85rem}
.btn:hover{border-color:var(--blue)}
.btn-primary{background:var(--blue);color:#fff;border-color:var(--blue)}
.btn-danger{color:var(--red);border-color:var(--red)}
.btn-row{display:flex;gap:8px;margin-top:16px;justify-content:flex-end}
.btn-status-close{padding:4px 12px;font-size:.75rem;border-color:var(--muted);color:var(--text2);flex-shrink:0;margin-left:8px}
.btn-status-close:hover{border-color:var(--text);color:var(--text)}
</style>
</head>
<body>

<div class="top-bar">
    <h1>&#9881; <span>Mantelroutine</span> — Schedule Matrix</h1>
    <div class="top-buttons">
        <button class="btn" onclick="triggerMantel()">&#9654; Mantel jetzt starten</button>
        <button class="btn btn-danger" onclick="shutdownServer()">&#9724; Schliessen und Stoppen</button>
    </div>
</div>

<!-- Live-Status-Banner -->
<div class="live-status" id="live-status">
    <span class="pulse" id="ls-pulse"></span>
    <div class="status-text">
        <div class="status-label" id="ls-label"></div>
        <div class="status-detail" id="ls-detail"></div>
        <div class="job-chips" id="ls-chips"></div>
    </div>
    <div class="status-timer" id="ls-timer-wrap" style="display:none">
        <div class="timer-value" id="ls-timer">0s</div>
        <div class="timer-label" id="ls-timer-label">Job laeuft</div>
    </div>
    <button class="btn btn-status-close" id="ls-close-btn" onclick="dismissStatus()" style="display:none">Schliessen</button>
</div>

<div class="last-run" id="last-run">
    <span class="dot" id="lr-dot"></span>
    <span id="lr-text">Lade...</span>
</div>

<div id="calendar"></div>

<!-- Edit Modal -->
<div class="modal-overlay" id="modal" onclick="if(event.target===this)closeModal()">
    <div class="modal">
        <h2>Aendern: <span id="modal-title"></span></h2>

        <div class="form-group">
            <label>Startzeit (HH:MM)</label>
            <input type="text" id="edit-startzeit" placeholder="03:00">
        </div>

        <div id="edit-jobs"></div>

        <div class="btn-row">
            <button class="btn" onclick="closeModal()">Abbrechen</button>
            <button class="btn btn-primary" onclick="saveDay()">Speichern</button>
        </div>
    </div>
</div>

<script>
var currentDatum = '';
var currentJobs = {};
var allDays = {};  // datum -> {startzeit, jobs, wochentag}
var statusPollTimer = null;
var triggerGraceUntil = 0;  // Timestamp: Polling nicht stoppen bei 'idle' bis dahin
var JOB_LABELS = {
    sec_delta_update: {name: 'SEC Delta Update', short: 'SEC', desc: 'SEC EDGAR Delta-Update: neue Filings laden'},
    aktien_delta_update: {name: 'Aktien Delta Update', short: 'Aktien', desc: 'Tagesaktuelle Kursdaten laden (yfinance)'},
    aktien_kalkulation: {name: 'Aktien Kalkulation', short: 'Kalk', desc: 'Berechnungen: Ampel, Breadth, Screening'}
};

async function loadData() {
    var resp = await fetch('/api/schedule');
    var data = await resp.json();
    renderLastRun(data.last_result);
    renderCalendar(data.weeks);
}

function renderLastRun(r) {
    var dot = document.getElementById('lr-dot');
    var text = document.getElementById('lr-text');
    if (!r || !r.date) { dot.className='dot empty'; text.textContent='Noch kein Lauf'; return; }
    dot.className = 'dot ' + (r.overall_status || 'empty');
    var msg = r.message || r.overall_status || '';
    text.textContent = 'Letzter Lauf: ' + r.date + ' ' + (r.startzeit_actual||'') + ' — ' + (r.schedule_file||'') + ' — ' + msg;
}

function renderCalendar(weeks) {
    allDays = {};
    var html = '';
    weeks.forEach(function(w) {
        html += '<div class="kw-label">' + w.label + '</div>';
        html += '<div class="week-grid">';
        w.days.forEach(function(d) {
            // Daten im JS-Objekt speichern statt inline
            allDays[d.datum] = {startzeit: d.startzeit, jobs: d.jobs, wochentag: d.wochentag};

            var cls = 'day-card';
            if (d.is_today) cls += ' today';
            if (d.is_past) cls += ' past';
            if (d.is_weekend) cls += ' weekend';

            var jobsHtml = '';
            var jobKeys = ['sec_delta_update', 'aktien_delta_update', 'aktien_kalkulation'];
            jobKeys.forEach(function(k) {
                var j = d.jobs && d.jobs[k] ? d.jobs[k] : {value:0, active:false, error:false};
                var label = JOB_LABELS[k] ? JOB_LABELS[k].name.split(' ')[0] : k;
                var dotCls = j.error ? 'error' : (j.active ? 'active' : 'inactive');
                var lblCls = j.error ? 'error' : (j.active ? 'active' : '');
                jobsHtml += '<div class="job-row-mini job-toggle-click" data-datum="' + d.datum + '" data-job="' + k + '" onclick="event.stopPropagation();toggleJob(this.dataset.datum,this.dataset.job)">';
                jobsHtml += '<span class="job-dot ' + dotCls + '"></span>';
                jobsHtml += '<span class="job-label ' + lblCls + '">' + label + '</span>';
                if (j.error) jobsHtml += '<span class="job-error-code">' + j.value + '</span>';
                jobsHtml += '</div>';
            });

            html += '<div class="' + cls + '" data-datum="' + d.datum + '" onclick="openModalFromData(this.dataset.datum)">';
            html += '<div class="day-name">' + d.wochentag_kurz + '</div>';
            html += '<div class="day-date">' + d.datum.substring(0, 5) + '</div>';
            html += '<div class="day-time">' + d.startzeit + '</div>';
            html += jobsHtml;
            html += '</div>';
        });
        html += '</div>';
    });
    document.getElementById('calendar').innerHTML = html;
}

function openModalFromData(datum) {
    var d = allDays[datum];
    if (!d) return;
    openModal(datum, d.wochentag + ' ' + datum, d.startzeit, d.jobs);
}

function openModal(datum, title, startzeit, jobs) {
    currentDatum = datum;
    currentJobs = {};
    for (var k in jobs) {
        currentJobs[k] = jobs[k].active || false;
    }
    document.getElementById('modal-title').textContent = title;
    document.getElementById('edit-startzeit').value = startzeit;

    var html = '';
    var jobKeys = ['sec_delta_update', 'aktien_delta_update', 'aktien_kalkulation'];
    jobKeys.forEach(function(k) {
        var info = JOB_LABELS[k] || {name: k, desc: ''};
        var j = jobs[k] || {active:false, error:false, value:0};
        var checked = j.active ? 'checked' : '';
        var statusTxt = j.error ? ' <span style="color:var(--red)">(Fehler: ' + j.value + ')</span>' : (j.active ? ' <span style="color:var(--green)">(OK)</span>' : '');
        html += '<div class="job-toggle">';
        html += '<input type="checkbox" id="job-' + k + '" ' + checked + ' onchange="currentJobs[\'' + k + '\']=this.checked">';
        html += '<div><div class="job-name">' + info.name + statusTxt + '</div><div class="job-desc">' + info.desc + '</div></div>';
        html += '</div>';
    });
    document.getElementById('edit-jobs').innerHTML = html;
    document.getElementById('modal').classList.add('open');
}

function closeModal() {
    document.getElementById('modal').classList.remove('open');
}

async function saveDay() {
    var startzeit = document.getElementById('edit-startzeit').value.trim() || '03:00';
    var resp = await fetch('/api/save', {
        method: 'POST',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({datum: currentDatum, startzeit: startzeit, jobs: currentJobs})
    });
    var r = await resp.json();
    if (r.ok) {
        closeModal();
        loadData();
    } else {
        alert('Speichern fehlgeschlagen');
    }
}

async function triggerMantel() {
    var resp = await fetch('/api/trigger', {method: 'POST'});
    var r = await resp.json();
    // Sofort Banner zeigen mit 'starting' Zustand
    triggerGraceUntil = Date.now() + 8000;  // 8 Sekunden Grace Period
    renderLiveStatus({state: 'starting'});
    startStatusPolling();
}

async function toggleJob(datum, jobKey) {
    var d = allDays[datum];
    if (!d) return;
    var newJobs = {};
    for (var k in d.jobs) {
        var j = d.jobs[k];
        if (k === jobKey) {
            // Toggle: aktiv/fehler -> aus, aus -> an
            newJobs[k] = j.active || j.error ? false : true;
        } else {
            newJobs[k] = j.active;
        }
    }
    var resp = await fetch('/api/save', {
        method: 'POST',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({datum: datum, startzeit: d.startzeit, jobs: newJobs})
    });
    var r = await resp.json();
    if (r.ok) loadData();
}

async function shutdownServer() {
    if (!confirm('Dashboard schliessen und Server stoppen?')) return;
    try { await fetch('/api/shutdown', {method:'POST'}); } catch(e) {}
    window.close();
}

// ======================================================
//  Live-Status-Polling
// ======================================================

function startStatusPolling() {
    if (statusPollTimer) return;  // laeuft bereits
    pollStatus();
    statusPollTimer = setInterval(pollStatus, 2000);
}

function stopStatusPolling() {
    if (statusPollTimer) {
        clearInterval(statusPollTimer);
        statusPollTimer = null;
    }
}

async function pollStatus() {
    try {
        var resp = await fetch('/api/status');
        var s = await resp.json();
        renderLiveStatus(s);
    } catch (e) {
        // Server weg?
    }
}

function renderLiveStatus(s) {
    var el = document.getElementById('live-status');
    var pulse = document.getElementById('ls-pulse');
    var label = document.getElementById('ls-label');
    var detail = document.getElementById('ls-detail');
    var chips = document.getElementById('ls-chips');
    var progressWrap = document.getElementById('ls-progress-wrap');
    var progressFill = document.getElementById('ls-progress');
    var progressText = document.getElementById('ls-progress-text');

    if (!s || s.state === 'idle') {
        // Grace Period nach Trigger: Mantel braucht 1-2s zum Starten
        if (Date.now() < triggerGraceUntil) return;  // Weiter pollen, nicht ausblenden
        el.classList.remove('visible');
        el.className = 'live-status';
        stopStatusPolling();
        return;
    }

    // Mantel hat sich gemeldet — Grace Period nicht mehr noetig
    triggerGraceUntil = 0;

    el.classList.add('visible');
    el.className = 'live-status visible state-' + s.state;

    if (s.state === 'starting') {
        pulse.className = 'pulse blue';
        label.textContent = 'Mantel wird gestartet...';
        detail.textContent = 'Warte auf Rueckmeldung';
        chips.innerHTML = '';
        progressWrap.style.display = 'none';
        document.getElementById('ls-close-btn').style.display = 'none';

    } else if (s.state === 'waiting') {
        pulse.className = 'pulse yellow';
        var mins = s.wait_seconds ? Math.round(s.wait_seconds / 60) : '?';
        label.textContent = 'Mantel wartet auf Startzeit: ' + (s.target_time || '--:--');
        detail.textContent = 'Noch ca. ' + mins + ' Minuten';
        chips.innerHTML = '';
        document.getElementById('ls-timer-wrap').style.display = 'none';

    } else if (s.state === 'running') {
        pulse.className = 'pulse blue';
        label.textContent = 'Job laeuft: ' + (s.current_job_beschreibung || s.current_job || '?');
        detail.textContent = 'Job ' + (s.job_index || '?') + ' von ' + (s.jobs_total || '?');

        // Elapsed-Timer: Job-Laufzeit + Mantel-Gesamtlaufzeit
        var timerWrap = document.getElementById('ls-timer-wrap');
        var timerEl = document.getElementById('ls-timer');
        var timerLabel = document.getElementById('ls-timer-label');
        timerWrap.style.display = '';
        if (s.job_started) {
            var jobSec = Math.round((Date.now() - new Date(s.job_started).getTime()) / 1000);
            timerEl.textContent = formatElapsed(jobSec);
        }
        // Mantel-Gesamtlaufzeit unter dem Job-Timer
        var mantelInfo = 'Job ' + (s.job_index || '') + '/' + (s.jobs_total || '');
        if (s.started) {
            var mantelSec = Math.round((Date.now() - new Date(s.started).getTime()) / 1000);
            mantelInfo += ' \u00b7 Mantel seit ' + formatElapsed(mantelSec);
        }
        timerLabel.textContent = mantelInfo;
        document.getElementById('ls-close-btn').style.display = 'none';

        // Job-Chips: done / active / pending
        var chipsHtml = '';
        var jobNames = s.job_names || [];
        var resultsSoFar = s.results_so_far || [];
        var doneNames = {};
        resultsSoFar.forEach(function(r) {
            doneNames[r.name] = r.status;
        });
        jobNames.forEach(function(name) {
            var shortName = JOB_LABELS[name] ? JOB_LABELS[name].short : name;
            if (doneNames[name]) {
                var cls = doneNames[name] === 'success' ? 'done' : 'error';
                chipsHtml += '<span class="job-chip ' + cls + '">' + shortName + '</span>';
            } else if (name === s.current_job) {
                chipsHtml += '<span class="job-chip active">' + shortName + '</span>';
            } else {
                chipsHtml += '<span class="job-chip pending">' + shortName + '</span>';
            }
        });
        chips.innerHTML = chipsHtml;

    } else if (s.state === 'finished') {
        pulse.className = 'pulse green';
        var ok = s.jobs_success || 0;
        var err = s.jobs_error || 0;
        label.textContent = 'Mantel abgeschlossen \u2014 ' + ok + ' OK' + (err > 0 ? ', ' + err + ' Fehler' : '');
        detail.textContent = 'Gesamtdauer: ' + formatElapsed(s.duration_sec || 0);
        document.getElementById('ls-timer-wrap').style.display = 'none';

        // Ergebnis-Chips
        var chipsHtml = '';
        (s.results || []).forEach(function(r) {
            var shortName = JOB_LABELS[r.name] ? JOB_LABELS[r.name].short : r.name;
            var cls = r.status === 'success' ? 'done' : 'error';
            var extra = r.error_code && r.error_code < 0 ? ' (' + r.error_code + ')' : '';
            chipsHtml += '<span class="job-chip ' + cls + '">' + shortName + extra + '</span>';
        });
        chips.innerHTML = chipsHtml;

        // Schliessen-Button zeigen, Polling stoppen, Kalender aktualisieren
        document.getElementById('ls-close-btn').style.display = '';
        stopStatusPolling();
        loadData();

    } else if (s.state === 'crashed') {
        pulse.className = 'pulse red';
        label.textContent = 'Mantel-Prozess abgestuerzt';
        detail.textContent = 'PID ' + (s.pid || '?') + ' antwortet nicht mehr';
        chips.innerHTML = '';
        progressWrap.style.display = 'none';

        // Schliessen-Button zeigen, Polling stoppen
        document.getElementById('ls-close-btn').style.display = '';
        stopStatusPolling();
    }
}

// Beim Start: einmal Status pruefen, dann Polling nur starten wenn Mantel aktiv
async function initStatusCheck() {
    try {
        var resp = await fetch('/api/status');
        var s = await resp.json();
        if (s.state && s.state !== 'idle') {
            startStatusPolling();
        }
    } catch (e) {}
}

function dismissStatus() {
    var el = document.getElementById('live-status');
    el.classList.remove('visible');
    el.className = 'live-status';
    document.getElementById('ls-close-btn').style.display = 'none';
    stopStatusPolling();
    loadData();
}

function formatElapsed(sec) {
    sec = Math.round(sec);
    if (sec < 60) return sec + 's';
    var m = Math.floor(sec / 60);
    var s = sec % 60;
    if (m < 60) return m + 'm ' + (s < 10 ? '0' : '') + s + 's';
    var h = Math.floor(m / 60);
    m = m % 60;
    return h + 'h ' + (m < 10 ? '0' : '') + m + 'm';
}

loadData();
initStatusCheck();
</script>
</body>
</html>""";


if __name__ == "__main__":
    os.chdir(SCRIPT_DIR)
    server = http.server.HTTPServer(("127.0.0.1", PORT), DashboardHandler)
    url = f"http://localhost:{PORT}"
    print(f"Morgenroutine Dashboard: {url}")
    print("Beenden: Ctrl+C oder Button im Dashboard")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nServer beendet.")
        server.server_close()
