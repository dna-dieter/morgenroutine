#!/usr/bin/env python3
"""
Simuliert Job-Ergebnisse in schedule_matrix.xlsx fuer Testzwecke.

Schreibt fuer den 02.04.2026 (heute):
  - SEC Delta Update:     -3  (Fehler, rot)
  - Aktien Delta Update:   1  (Erfolg, gruen)
  - Aktien Kalkulation:   -12 (Fehler, rot)

Aufruf: python3 simulate_results.py
Danach Dashboard neu laden (F5) um das Ergebnis zu sehen.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from pathlib import Path

XLSX = Path(__file__).parent / "schedule_matrix.xlsx"

GREEN_BG = PatternFill("solid", fgColor="1A3A2A")
GREEN_FONT = Font(name="Arial", size=11, bold=True, color="3FB950")
RED_BG = PatternFill("solid", fgColor="3A1A1A")
RED_FONT = Font(name="Arial", size=11, bold=True, color="F85149")
DIM_BG = PatternFill("solid", fgColor="0D1117")
DIM_FONT = Font(name="Arial", size=10, color="30363D")

# Simulation: Datum -> {Spalte: Wert}
# Spalte 5=SEC(1), 7=Aktien(1), 9=Kalk(1)  (1-Zellen)
# Spalte 4=SEC(0), 6=Aktien(0), 8=Kalk(0)  (0-Zellen)
SIMULATIONS = {
    "02.04.2026": {
        5: -3,    # SEC: Fehler -03
        7: 1,     # Aktien Delta: Erfolg
        9: -12,   # Kalkulation: Fehler -12
    }
}

wb = load_workbook(XLSX)
ws = wb.active

for row_idx in range(3, ws.max_row + 1):
    cell_val = ws.cell(row=row_idx, column=1).value
    if cell_val and str(cell_val).strip() in SIMULATIONS:
        datum = str(cell_val).strip()
        sim = SIMULATIONS[datum]
        print(f"\n{datum}:")

        for col, value in sim.items():
            cell = ws.cell(row=row_idx, column=col)
            cell.value = value

            # 0-Zelle (links) leeren
            col_0 = col - 1
            c0 = ws.cell(row=row_idx, column=col_0)
            c0.value = None
            c0.font = DIM_FONT
            c0.fill = DIM_BG

            if value > 0:
                cell.font = GREEN_FONT
                cell.fill = GREEN_BG
                print(f"  Spalte {col}: +{value} (Erfolg)")
            elif value < 0:
                cell.font = RED_FONT
                cell.fill = RED_BG
                print(f"  Spalte {col}: {value} (FEHLER)")

wb.save(XLSX)
print(f"\nGespeichert: {XLSX}")
print("Dashboard neu laden (F5) um Ergebnis zu sehen.")
